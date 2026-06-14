from datetime import date, datetime

from openpyxl import load_workbook

from app.scanners.document_blocks import ImportedBlock, clean_text, rows_to_text


def scan_xlsx_to_blocks(file_path):
    workbook = load_workbook(str(file_path), read_only=True, data_only=True)
    blocks = []

    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = []

            for row in worksheet.iter_rows(values_only=True):
                values = [_format_cell_value(value) for value in row]

                while values and not values[-1]:
                    values.pop()

                if any(values):
                    rows.append(values)

            if rows:
                blocks.append(
                    ImportedBlock(
                        block_type="spreadsheet_table",
                        content_text=rows_to_text(rows),
                        content_json={
                            "sheet_name": sheet_name,
                            "rows": rows,
                        },
                        order_index=len(blocks),
                        rows=rows,
                    )
                )
    finally:
        workbook.close()

    if not blocks:
        raise ValueError("Excel-файл не содержит данных для импорта.")

    return blocks


def _format_cell_value(value):
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")

    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")

    return clean_text(value)

