from datetime import date, datetime

from openpyxl import load_workbook

from app.base_scanner import (
    clean_cell_value,
    extract_key_values_from_table,
    finalize_scan_result,
    make_empty_scan_result,
)


def scan_xlsx(file_path):
    workbook = load_workbook(str(file_path), read_only=True, data_only=True)
    result = make_empty_scan_result("xlsx")
    result["sheets_count"] = len(workbook.sheetnames)

    text_parts = []

    try:
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            worksheet = workbook[sheet_name]
            sheet_rows = []

            for row in worksheet.iter_rows(values_only=True):
                values = [_format_cell_value(value) for value in row]

                while values and not values[-1]:
                    values.pop()

                if not any(values):
                    continue

                sheet_rows.append(values)
                text_parts.append(" | ".join(value for value in values if value))

            if sheet_rows:
                table = {
                    "sheet_name": sheet_name,
                    "rows": sheet_rows,
                }
                result["tables"].append(table)
                result["key_values"].update(
                    extract_key_values_from_table(sheet_rows, sheet_index)
                )
    finally:
        workbook.close()

    result["plain_text"] = "\n".join(text_parts).strip()

    if not result["plain_text"]:
        raise ValueError("XLSX не содержит данных для импорта.")

    return finalize_scan_result(result)


def _format_cell_value(value):
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")

    return clean_cell_value(value)
